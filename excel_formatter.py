import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os
import pandas as pd

def get_short_name(full_header):
    """Extract the short name from the full parameter description."""
    if full_header == "Document":
        return "Document"
    
    if "**Parameter:**" in full_header:
        param_text = full_header.split("**Parameter:**")[1].split("\n")[0].strip()
        return param_text.rstrip('.')
    
    return full_header

def get_column_mapping(df):
    """Create a mapping between desired column names and their indices in the DataFrame."""
    mapping = {}
    for idx, col in enumerate(df.columns):
        if col == "Document":
            mapping["Document"] = idx
        elif "CGIAR Region" in col:
            mapping["CGIAR Region"] = idx
        elif "Category for Natural Resource Management" in col:
            mapping["Category for Natural Resource Management"] = idx
        elif "Category of Study" in col:
            mapping["Category of Study"] = idx
        elif "Contributing Initiatives or centers" in col:
            mapping["Contributing Initiatives or centers"] = idx
        elif "Country of Study" in col:
            mapping["Country of Study"] = idx
        elif "Crops" in col:
            mapping["Crops"] = idx
        elif "Keywords" in col:
            mapping["Keywords"] = idx
        elif "Link or DOI" in col:
            mapping["Link or DOI"] = idx
        elif "Name; Unit/Metric; and Value for Reported Indicator of Impact" in col:
            mapping["Name; Unit/Metric; and Value for Reported Indicator of Impact"] = idx
        elif "Period Analyzed" in col:
            mapping["Period Analyzed"] = idx
        elif "Primary Product Type" in col:
            mapping["Primary Product Type"] = idx
        elif "Primary and Secondary CGIAR Impact Area(s)" in col:
            mapping["Primary and Secondary CGIAR Impact Area(s)"] = idx
        elif "Study Theme" in col:
            mapping["Study Theme"] = idx
        elif "Summary" in col:
            mapping["Summary"] = idx
        elif "Title" in col:
            mapping["Title"] = idx
        elif "Year of Report" in col:
            mapping["Year of Report"] = idx
    return mapping

def format_excel():
    # Load the workbook
    current_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(current_dir, "output_single_question_pivoted.xlsx")
    output_file = os.path.join(current_dir, "output_single_question_pivoted_formatted_new.xlsx")
    
    # Read with pandas first to get the actual column names
    df = pd.read_excel(input_file)
    print("Original columns:", list(df.columns))
    
    # Define the desired column order
    column_order = [
        # Basic Information
        'Document',
        'Title',
        'Year of Report',
        'Summary',
        
        # Study Details
        'Study Theme',
        'Category of Study',
        'Period Analyzed',
        'Link or DOI',
        
        # Geographic & Institutional Info
        'Country of Study',
        'CGIAR Region',
        'Contributing Initiatives or centers',
        
        # Subject Matter
        'Primary Product Type',
        'Crops',
        'Category for Natural Resource Management',
        'Keywords',
        
        # Impact Information
        'Primary and Secondary CGIAR Impact Area(s)',
        'Name; Unit/Metric; and Value for Reported Indicator of Impact'
    ]
    
    # Get column mapping
    col_mapping = get_column_mapping(df)
    print("Column mapping:", col_mapping)
    
    # Create a new DataFrame with the desired column order
    new_df = pd.DataFrame(columns=column_order)
    
    # Copy data to the new DataFrame
    for col in column_order:
        if col in col_mapping:
            new_df[col] = df.iloc[:, col_mapping[col]]
    
    # Sort the DataFrame by Document column
    new_df = new_df.sort_values('Document')
    
    # Print first few rows of data for verification
    print("\nFirst few rows of data (after sorting):")
    print(new_df.head())
    
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="004F9F", end_color="004F9F", fill_type="solid")  # CGIAR blue
    header_font = Font(color="FFFFFF", bold=True)
    alternate_row_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side('thin')
    )
    
    # Alignments
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    long_text_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Write headers
    for col_idx, col_name in enumerate(column_order, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Write data and apply formatting
    for row_idx, row in enumerate(new_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Apply alternating row colors
            if row_idx % 2 == 0:
                cell.fill = alternate_row_fill
            
            # Apply specific column formatting
            col_name = column_order[col_idx - 1]
            if col_name in ['Summary', 'Study Theme', 'Title']:
                cell.alignment = long_text_alignment
                ws.column_dimensions[get_column_letter(col_idx)].width = 50
            elif col_name == 'Document':
                cell.alignment = left_alignment
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
            elif col_name == 'Link or DOI':
                cell.alignment = left_alignment
                ws.column_dimensions[get_column_letter(col_idx)].width = 40
            elif col_name == 'Year of Report':
                cell.alignment = center_alignment
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            else:
                cell.alignment = left_alignment
                ws.column_dimensions[get_column_letter(col_idx)].width = 30
    
    # Set row heights
    ws.row_dimensions[1].height = 45  # Header row
    for row in range(2, len(new_df) + 2):
        ws.row_dimensions[row].height = 25  # Data rows
    
    # Save the workbook
    wb.save(output_file)
    print(f"\nFormatted Excel file saved as: {output_file}")

if __name__ == "__main__":
    format_excel() 